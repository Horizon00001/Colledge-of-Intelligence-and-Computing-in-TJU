from openpyxl import load_workbook
from pathlib import Path
wb = load_workbook(str(Path(__file__).parent / 'cic.xlsx'))
ws = wb.active
print(str(Path(__file__).parent / 'cic.xlsx'))
column_ID = [cell.value for cell in ws['A']]
column_Name = [cell.value for cell in ws['B']]
print(column_ID)
